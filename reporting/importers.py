"""Parsing and import of admin-uploaded facility, program and data files."""

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from .models import DataRecord, Facility, Indicator, Program, ReportingPeriod, UploadBatch

CSV_EXTENSIONS = (".csv",)
EXCEL_EXTENSIONS = (".xlsx", ".xlsm")

FACILITY_COLUMNS = ["code", "name"]
PROGRAM_COLUMNS = ["code", "name"]
DATA_COLUMNS = [
    "facility_code",
    "program_code",
    "indicator_code",
    "indicator_name",
    "period_code",
    "value",
]

TEMPLATE_HEADERS = {
    UploadBatch.FACILITIES: ["code", "name", "district", "facility_type", "is_active"],
    UploadBatch.PROGRAMS: ["code", "name", "description", "is_active"],
    UploadBatch.DATA: [
        "facility_code",
        "program_code",
        "indicator_code",
        "indicator_name",
        "indicator_unit",
        "period_code",
        "period_name",
        "period_type",
        "period_start",
        "period_end",
        "value",
    ],
}

TRUE_VALUES = {"1", "true", "yes", "y", "t", "active"}
FALSE_VALUES = {"0", "false", "no", "n", "f", "inactive"}
DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d")


class ImportError_(Exception):
    """Raised when an uploaded file cannot be parsed at all."""


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def _normalise_header(value) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_upload(uploaded_file) -> list[dict[str, str]]:
    """Parse a CSV or Excel upload into a list of lower-cased-key row dicts."""
    name = (getattr(uploaded_file, "name", "") or "").lower()
    content = uploaded_file.read()
    if isinstance(content, str):
        content = content.encode("utf-8")

    if name.endswith(EXCEL_EXTENSIONS):
        rows = _parse_excel(content)
    elif name.endswith(CSV_EXTENSIONS):
        rows = _parse_csv(content)
    else:
        raise ImportError_("Unsupported file type. Upload a .csv or .xlsx file.")

    if not rows:
        raise ImportError_("The uploaded file is empty.")
    return rows


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        raise ImportError_("The uploaded file is empty.")
    headers = [_normalise_header(cell) for cell in header]
    rows = []
    for raw in reader:
        if not any(str(cell).strip() for cell in raw):
            continue
        rows.append({h: str(v).strip() for h, v in zip(headers, raw) if h})
    return rows


def _parse_excel(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ImportError_("The uploaded file is empty.")
    headers = [_normalise_header(cell) for cell in header]
    rows = []
    for raw in rows_iter:
        if not any(_cell_to_text(cell) for cell in raw):
            continue
        rows.append({h: _cell_to_text(v) for h, v in zip(headers, raw) if h})
    workbook.close()
    return rows


def _missing_columns(rows: list[dict[str, str]], required: list[str]) -> list[str]:
    present = set(rows[0].keys())
    return [column for column in required if column not in present]


def _parse_bool(value: str, default: bool = True) -> bool:
    text = (value or "").strip().lower()
    if not text:
        return default
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    raise ValueError(f"'{value}' is not a valid true/false value")


def _parse_date(value: str) -> date:
    text = (value or "").strip()
    if not text:
        raise ValueError("date is required")
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"'{value}' is not a valid date (use YYYY-MM-DD)")


def _parse_decimal(value: str) -> Decimal:
    text = (value or "").strip().replace(",", "")
    if not text:
        raise ValueError("value is required")
    try:
        return Decimal(text)
    except InvalidOperation:
        raise ValueError(f"'{value}' is not a number")


def import_facilities(rows: list[dict[str, str]]) -> ImportResult:
    result = ImportResult()
    missing = _missing_columns(rows, FACILITY_COLUMNS)
    if missing:
        result.errors.append(f"Missing required column(s): {', '.join(missing)}")
        return result

    for line, row in enumerate(rows, start=2):
        code = row.get("code", "").strip()
        name = row.get("name", "").strip()
        if not code or not name:
            result.errors.append(f"Row {line}: 'code' and 'name' are required")
            continue
        try:
            is_active = _parse_bool(row.get("is_active", ""))
        except ValueError as exc:
            result.errors.append(f"Row {line}: {exc}")
            continue
        _, created = Facility.objects.update_or_create(
            code=code,
            defaults={
                "name": name,
                "district": row.get("district", "").strip(),
                "facility_type": row.get("facility_type", "").strip(),
                "is_active": is_active,
            },
        )
        if created:
            result.created += 1
        else:
            result.updated += 1
    return result


def import_programs(rows: list[dict[str, str]]) -> ImportResult:
    result = ImportResult()
    missing = _missing_columns(rows, PROGRAM_COLUMNS)
    if missing:
        result.errors.append(f"Missing required column(s): {', '.join(missing)}")
        return result

    for line, row in enumerate(rows, start=2):
        code = row.get("code", "").strip()
        name = row.get("name", "").strip()
        if not code or not name:
            result.errors.append(f"Row {line}: 'code' and 'name' are required")
            continue
        try:
            is_active = _parse_bool(row.get("is_active", ""))
        except ValueError as exc:
            result.errors.append(f"Row {line}: {exc}")
            continue
        _, created = Program.objects.update_or_create(
            code=code,
            defaults={
                "name": name,
                "description": row.get("description", "").strip(),
                "is_active": is_active,
            },
        )
        if created:
            result.created += 1
        else:
            result.updated += 1
    return result


def _get_or_create_period(row: dict[str, str], line: int) -> ReportingPeriod:
    code = row.get("period_code", "").strip()
    period = ReportingPeriod.objects.filter(code=code).first()
    if period:
        return period
    start = _parse_date(row.get("period_start", ""))
    end = _parse_date(row.get("period_end", ""))
    if start > end:
        raise ValueError("period_start must not be after period_end")
    period_type = (row.get("period_type", "") or ReportingPeriod.MONTHLY).strip().lower()
    valid_types = {choice for choice, _ in ReportingPeriod.PERIOD_TYPES}
    if period_type not in valid_types:
        raise ValueError(f"'{period_type}' is not a valid period_type ({', '.join(sorted(valid_types))})")
    return ReportingPeriod.objects.create(
        code=code,
        name=row.get("period_name", "").strip() or code,
        period_type=period_type,
        start_date=start,
        end_date=end,
    )


def import_data_records(rows: list[dict[str, str]], batch: UploadBatch | None = None) -> ImportResult:
    """Import reported values, creating indicators and periods described in the file.

    Facilities and programs must already exist so that a typo in a code is reported
    instead of silently creating a duplicate facility.
    """
    result = ImportResult()
    missing = _missing_columns(rows, DATA_COLUMNS)
    if missing:
        result.errors.append(f"Missing required column(s): {', '.join(missing)}")
        return result

    facilities = {facility.code: facility for facility in Facility.objects.all()}
    programs = {program.code: program for program in Program.objects.all()}

    for line, row in enumerate(rows, start=2):
        facility = facilities.get(row.get("facility_code", "").strip())
        program = programs.get(row.get("program_code", "").strip())
        if facility is None:
            result.errors.append(f"Row {line}: unknown facility_code '{row.get('facility_code', '')}'")
            continue
        if program is None:
            result.errors.append(f"Row {line}: unknown program_code '{row.get('program_code', '')}'")
            continue

        indicator_code = row.get("indicator_code", "").strip()
        period_code = row.get("period_code", "").strip()
        if not indicator_code or not period_code:
            result.errors.append(f"Row {line}: 'indicator_code' and 'period_code' are required")
            continue

        try:
            value = _parse_decimal(row.get("value", ""))
            period = _get_or_create_period(row, line)
        except ValueError as exc:
            result.errors.append(f"Row {line}: {exc}")
            continue

        indicator, _ = Indicator.objects.get_or_create(
            program=program,
            code=indicator_code,
            defaults={
                "name": row.get("indicator_name", "").strip() or indicator_code,
                "unit": row.get("indicator_unit", "").strip(),
            },
        )

        _, created = DataRecord.objects.update_or_create(
            facility=facility,
            program=program,
            indicator=indicator,
            period=period,
            defaults={"value": value, "batch": batch},
        )
        if created:
            result.created += 1
        else:
            result.updated += 1
    return result


IMPORTERS = {
    UploadBatch.FACILITIES: import_facilities,
    UploadBatch.PROGRAMS: import_programs,
}


def run_import(dataset: str, uploaded_file, user=None) -> tuple[UploadBatch, ImportResult]:
    """Parse and import an upload atomically; nothing is saved if any row fails."""
    rows = parse_upload(uploaded_file)
    batch = UploadBatch(
        dataset=dataset,
        file_name=getattr(uploaded_file, "name", "upload"),
        uploaded_by=user if user is not None and user.is_authenticated else None,
    )

    try:
        with transaction.atomic():
            batch.save()
            if dataset == UploadBatch.DATA:
                result = import_data_records(rows, batch=batch)
            else:
                result = IMPORTERS[dataset](rows)
            if not result.ok:
                raise _Rollback(result)
    except _Rollback as rollback:
        result = rollback.result
        batch.pk = None
        batch.status = UploadBatch.FAILED
        batch.error_log = "\n".join(result.errors)
        batch.save()
        return batch, result

    batch.status = UploadBatch.COMPLETED
    batch.rows_created = result.created
    batch.rows_updated = result.updated
    batch.save(update_fields=["status", "rows_created", "rows_updated", "updated_at"])
    return batch, result


class _Rollback(Exception):
    def __init__(self, result: ImportResult):
        super().__init__("import failed")
        self.result = result
